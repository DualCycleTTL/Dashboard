import io
import re
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font

st.set_page_config(page_title="Dashboard Analisis Dual Cycle", layout="wide")

# ================================================================
# KONSTANTA DEFAULT (samakan dengan VBA)
# ================================================================
AMBANG_COMBO_MENIT_DEFAULT = 40
AMBANG_DUAL_MENIT_DEFAULT = 240  # = 4 jam (bukan 6 jam spt komentar VBA lama)
AMBANG_TWINLIFT_MENIT_DEFAULT = 1  # jarak DISC_LOAD_TS antar 2 kontainer combo

# Ukuran kontainer eligible utk Combo di VBA di-hardcode "= 20" (bukan
# pengaturan yang bisa diubah user), jadi di sini juga dikunci tetap 20ft
# supaya hasil selalu identik dengan VBA.
SIZE_ELIGIBLE = 20


# ================================================================
# HELPER - baca & siapkan data
# ================================================================

def klasifikasi_activity(val: object) -> str:
    s = str(val).upper()
    return "LOAD" if "LOAD" in s else "DISC"


_VBA_VAL_RE = re.compile(r"^\s*[+-]?\d+(\.\d+)?")


def vba_val(x: object) -> float:
    """
    Replikasi fungsi Val() di VBA: baca angka dari AWAL string sampai
    ketemu karakter non-angka pertama, sisanya diabaikan (mis. "20FT"
    -> 20, "  -5.5kg" -> -5.5). Kalau tidak ada angka sama sekali -> 0.
    Ini supaya CTR_SIZE yang formatnya campur teks (mis. "20FT") tetap
    dibaca sama seperti macro VBA aslinya, bukan malah jadi 0 seperti
    kalau pakai pd.to_numeric biasa.
    """
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)
    m = _VBA_VAL_RE.match(str(x))
    return float(m.group()) if m else 0.0


@st.cache_data(show_spinner=False, max_entries=3, ttl=1800)
def baca_file(file_bytes: bytes, filename: str, sheet_name=None):
    if filename.lower().endswith(".csv"):
        return {"__csv__": pd.read_csv(BytesIO(file_bytes))}
    xls = pd.ExcelFile(BytesIO(file_bytes))
    if sheet_name is not None:
        return {sheet_name: xls.parse(sheet_name)}
    return {name: xls.parse(name) for name in xls.sheet_names}


def siapkan_data(raw: pd.DataFrame, col_map: dict, size_eligible: int) -> pd.DataFrame:
    df = pd.DataFrame()
    df["VES_ID"] = raw[col_map["ves_id"]]
    df["CTR_SIZE"] = raw[col_map["size"]].apply(vba_val)
    df["CAR_CHE_ID"] = raw[col_map["truck"]].astype(str).str.strip()
    df["ACTIVITY"] = raw[col_map["activity"]].apply(klasifikasi_activity)
    df["TS_G"] = pd.to_datetime(raw[col_map["ts_g"]], errors="coerce")
    df["TS_H"] = pd.to_datetime(raw[col_map["ts_h"]], errors="coerce")

    # kalau G kosong pakai H, kalau H kosong pakai G (spt VBA)
    both_invalid = df["TS_G"].isna() & df["TS_H"].isna()
    df["TS_G"] = df["TS_G"].fillna(df["TS_H"])
    df["TS_H"] = df["TS_H"].fillna(df["TS_G"])

    # ------------------------------------------------------------
    # PENTING: VBA TIDAK PERNAH MEMBUANG BARIS, walau kedua kolom
    # timestamp kosong. Baris begitu tetap diproses sbg 1 event
    # tersendiri dengan tanggal dummy CDate(-1) = 29 Des 1899.
    # Versi sebelumnya di app ini MEMBUANG baris tsb (dropna),
    # sehingga Total Event & persentase jadi beda dari hasil VBA
    # (mis. 62,5% vs 63,2%). Sekarang disamakan persis dgn VBA:
    # baris tetap dihitung, bukan dibuang.
    # ------------------------------------------------------------
    if both_invalid.any():
        dummy_ts = pd.Timestamp("1899-12-29")  # = CDate(-1) di VBA
        df.loc[both_invalid, "TS_G"] = dummy_ts
        df.loc[both_invalid, "TS_H"] = dummy_ts
        st.warning(
            f"{int(both_invalid.sum())} baris punya kedua kolom timestamp "
            f"(DISC_LOAD_TS & STACK_UNSTACK_TS) kosong/tidak valid. Mengikuti "
            f"perilaku VBA, baris ini TETAP diproses sbg event tersendiri "
            f"(tanggal dummy 29 Des 1899), bukan dibuang, supaya total & "
            f"persentase persis sama dengan hasil macro VBA."
        )

    df = df.reset_index(drop=True)
    df["ROW_IDX"] = df.index
    return df


# ================================================================
# LAYER 1 - COMBO / SINGLE
# ================================================================

def layer1_combo(df: pd.DataFrame, ambang_combo: float, size_eligible: int) -> pd.DataFrame:
    n = len(df)
    ts_g = df["TS_G"].to_numpy()
    ts_h = df["TS_H"].to_numpy()
    size = df["CTR_SIZE"].to_numpy()
    activity = df["ACTIVITY"].to_numpy()

    assigned = np.zeros(n, dtype=bool)
    group_id = np.zeros(n, dtype=int)

    pairs = []
    truck_positions = df.groupby("CAR_CHE_ID").indices  # dict truk -> array posisi (0..n-1)

    thr_delta = np.timedelta64(int(round(ambang_combo * 60)), "s")

    for _, pos in truck_positions.items():
        for act in ("LOAD", "DISC"):
            idx_act = np.array([p for p in pos if activity[p] == act and size[p] == size_eligible])
            m = len(idx_act)
            if m < 2:
                continue

            # ------------------------------------------------------
            # min(gapG, gapH) <= ambang  <=>  gapG <= ambang ATAU
            # gapH <= ambang. Jadi kandidat cukup dicari lewat 2
            # sliding-window (satu berdasar TS_G, satu TS_H) yang
            # jauh lebih cepat (O(m log m)) daripada cek SEMUA
            # pasangan (O(m^2)) -- hasil akhirnya identik.
            # ------------------------------------------------------
            found = set()

            for ts_arr in (ts_g, ts_h):
                order = idx_act[np.argsort(ts_arr[idx_act])]
                sorted_ts = ts_arr[order]
                left = 0
                for right in range(len(order)):
                    while sorted_ts[right] - sorted_ts[left] > thr_delta:
                        left += 1
                    for b in range(left, right):
                        i, k = order[b], order[right]
                        if i > k:
                            i, k = k, i
                        found.add((int(i), int(k)))

            for i, k in found:
                gap_g = abs((ts_g[k] - ts_g[i]) / np.timedelta64(1, "m"))
                gap_h = abs((ts_h[k] - ts_h[i]) / np.timedelta64(1, "m"))
                gap = min(gap_g, gap_h)
                pairs.append((i, k, gap))

    # urutkan gap terkecil dulu, lalu greedy matching
    pairs.sort(key=lambda x: (x[2], x[0], x[1]))

    nxt = 0
    for i, k, _gap in pairs:
        if not assigned[i] and not assigned[k]:
            nxt += 1
            group_id[i] = nxt
            group_id[k] = nxt
            assigned[i] = assigned[k] = True

    for i in range(n):
        if not assigned[i]:
            nxt += 1
            group_id[i] = nxt
            assigned[i] = True

    out = df.copy()
    out["GROUP_ID"] = group_id
    return out


# ================================================================
# LAYER 1b - TWINLIFT (sub-analisis di dalam grup Combo)
# ================================================================

def deteksi_twinlift(df_combo: pd.DataFrame, ambang_twinlift: float, size_eligible: int) -> dict:
    """
    Twinlift adalah kondisi khusus DI DALAM grup Combo (2 baris jadi 1
    event, lihat layer1_combo) yang memenuhi SEMUA syarat berikut:
      1) Muatan Combo 20ft -- otomatis terpenuhi krn Combo di layer1
         memang hanya dibentuk dari baris dengan CTR_SIZE == size_eligible
         (20ft), tapi tetap dicek ulang di sini supaya aman/eksplisit.
      2) Berasal dari kapal yang sama -> VES_ID kedua baris identik.
      3) Waktu DISC_LOAD_TS (TS_G) antar kedua baris berdekatan, yaitu
         selisihnya <= ambang_twinlift menit (default 1 menit).

    Grup Single (cuma 1 anggota, tidak ber-Combo) otomatis BUKAN
    Twinlift, ditandai "-" (bukan "Bukan Twinlift") supaya gampang
    dibedakan dari Combo yang gagal syarat Twinlift.

    Mengembalikan dict {GROUP_ID: status} dengan status salah satu dari
    "Twinlift", "Bukan Twinlift", atau "-" (utk grup Single).
    """
    result = {}
    for gid, g in df_combo.groupby("GROUP_ID"):
        if len(g) != 2:
            result[gid] = "-"
            continue

        g = g.sort_values("ROW_IDX")
        r1, r2 = g.iloc[0], g.iloc[1]

        syarat_size = (r1["CTR_SIZE"] == size_eligible) and (r2["CTR_SIZE"] == size_eligible)
        syarat_kapal = r1["VES_ID"] == r2["VES_ID"]
        gap_disc_load = abs((r2["TS_G"] - r1["TS_G"]) / np.timedelta64(1, "m"))
        syarat_waktu = gap_disc_load <= ambang_twinlift

        if syarat_size and syarat_kapal and syarat_waktu:
            result[gid] = "Twinlift"
        else:
            result[gid] = "Bukan Twinlift"

    return result


# ================================================================
# BENTUK EVENT
# ================================================================

def bentuk_event(df: pd.DataFrame) -> pd.DataFrame:
    # Vektorisasi penuh (bukan loop per grup) supaya tetap cepat
    # meski jumlah grup/event sangat banyak.
    is_disc = df["ACTIVITY"].to_numpy() == "DISC"
    evt_start = np.where(is_disc, df["TS_G"].to_numpy(), df["TS_H"].to_numpy())
    evt_end = np.where(is_disc, df["TS_H"].to_numpy(), df["TS_G"].to_numpy())

    tmp = pd.DataFrame(
        {
            "GROUP_ID": df["GROUP_ID"].to_numpy(),
            "ACTIVITY": df["ACTIVITY"].to_numpy(),
            "CAR_CHE_ID": df["CAR_CHE_ID"].to_numpy(),
            "EVT_START": evt_start,
            "EVT_END": evt_end,
        }
    )

    # PENTING: groupby harus sort=True (urut menaik berdasarkan GROUP_ID),
    # BUKAN sort=False (yang ikut urutan kemunculan baris di data asli).
    # Ini supaya "posisi ke-0,1,2,..." event di sini persis sama dengan
    # urutan "g = 1 To nGrp" di VBA -- karena posisi ini dipakai sbg
    # tie-breaker saat sorting kandidat pasangan Dual Cycle yang gap-nya
    # SAMA PERSIS (sering terjadi krn banyak event overlap = gap 0 menit).
    # Kalau urutannya beda dari VBA, greedy matching-nya bisa memilih
    # pasangan yang berbeda meski total event-nya tetap sama -- inilah
    # yang menyebabkan Dual Cycle vs Non Dual beda antara VBA & dashboard
    # padahal Total Event-nya sudah identik.
    events = tmp.groupby("GROUP_ID", sort=True).agg(
        ACTIVITY=("ACTIVITY", "first"),
        CAR_CHE_ID=("CAR_CHE_ID", "first"),
        START_TS=("EVT_START", "min"),
        END_TS=("EVT_END", "max"),
        N_ANGGOTA=("GROUP_ID", "size"),
    ).reset_index()

    events["CONTAINER_STATUS"] = np.where(events["N_ANGGOTA"] >= 2, "Combo", "Single")
    events = events.drop(columns=["N_ANGGOTA"])
    return events


# ================================================================
# LAYER 2 - DUAL CYCLE
# ================================================================

def layer2_dual(events: pd.DataFrame, ambang_dual: float) -> pd.DataFrame:
    events = events.reset_index(drop=True)
    n = len(events)
    start = events["START_TS"].to_numpy()
    end = events["END_TS"].to_numpy()
    activity = events["ACTIVITY"].to_numpy()

    assigned = np.zeros(n, dtype=bool)
    status = np.array(["Non Dual"] * n, dtype=object)

    pairs = []
    truck_positions = events.groupby("CAR_CHE_ID").indices

    for _, pos in truck_positions.items():
        pos_sorted = sorted(pos, key=lambda p: start[p])
        m = len(pos_sorted)
        for a in range(m - 1):
            i = pos_sorted[a]
            for b in range(a + 1, m):
                k = pos_sorted[b]
                gap_ab = (start[k] - end[i]) / np.timedelta64(1, "m")
                gap_ba = (start[i] - end[k]) / np.timedelta64(1, "m")
                if gap_ab >= 0:
                    gap = gap_ab
                elif gap_ba >= 0:
                    gap = gap_ba
                else:
                    gap = 0.0  # overlap waktu -> dianggap berdekatan

                if gap > ambang_dual:
                    # waktu mulai naik monoton -> aman berhenti di sini
                    break

                if activity[i] != activity[k]:
                    pairs.append((i, k, gap))

    pairs.sort(key=lambda x: (x[2], x[0], x[1]))
    for i, k, _gap in pairs:
        if not assigned[i] and not assigned[k]:
            status[i] = "Dual Cycle"
            status[k] = "Dual Cycle"
            assigned[i] = assigned[k] = True

    out = events.copy()
    out["STATUS"] = status
    return out


# ================================================================
# PENOMORAN EVENT_ID GLOBAL (urut truk sesuai kemunculan pertama
# di data asli, di dalam truk diurutkan berdasarkan START_TS)
# ================================================================

def beri_event_id(events: pd.DataFrame, df_asli: pd.DataFrame):
    truck_order = list(dict.fromkeys(df_asli["CAR_CHE_ID"].tolist()))
    rank = {tk: i for i, tk in enumerate(truck_order)}

    events = events.copy()
    events["_truck_rank"] = events["CAR_CHE_ID"].map(rank)
    events = events.sort_values(["_truck_rank", "START_TS"]).reset_index(drop=True)
    events["EVENT_ID"] = events.index + 1
    events = events.drop(columns=["_truck_rank"])

    event_id_map = dict(zip(events["GROUP_ID"], events["EVENT_ID"]))
    return events, event_id_map


def gabungkan_hasil(df: pd.DataFrame, events: pd.DataFrame, event_id_map: dict) -> pd.DataFrame:
    status_map = events.set_index("GROUP_ID")["STATUS"].to_dict()
    container_map = events.set_index("GROUP_ID")["CONTAINER_STATUS"].to_dict()
    twinlift_map = events.set_index("GROUP_ID")["TWINLIFT_STATUS"].to_dict()

    out = df.copy()
    out["EVENT_ID"] = out["GROUP_ID"].map(event_id_map)
    out["CONTAINER_STATUS"] = out["GROUP_ID"].map(container_map)
    out["STATUS"] = out["GROUP_ID"].map(status_map)
    out["TWINLIFT_STATUS"] = out["GROUP_ID"].map(twinlift_map)
    out = out.drop(columns=["GROUP_ID", "ROW_IDX"])
    return out


# ================================================================
# RINGKASAN / STATISTIK
# ================================================================

def hitung_ringkasan(events: pd.DataFrame, out_df: pd.DataFrame) -> dict:
    total_event = len(events)
    total_dual = int((events["STATUS"] == "Dual Cycle").sum())
    total_single = total_event - total_dual

    combo_dual = int(((events["CONTAINER_STATUS"] == "Combo") & (events["STATUS"] == "Dual Cycle")).sum())
    combo_single = int(((events["CONTAINER_STATUS"] == "Combo") & (events["STATUS"] == "Non Dual")).sum())
    single_dual = int(((events["CONTAINER_STATUS"] == "Single") & (events["STATUS"] == "Dual Cycle")).sum())
    single_single = int(((events["CONTAINER_STATUS"] == "Single") & (events["STATUS"] == "Non Dual")).sum())

    # ------------------------------------------------------
    # TWINLIFT -- sub-analisis di dalam grup Combo. Basisnya event
    # (bukan baris), sama seperti Combo/Single & Dual Cycle/Non Dual.
    # ------------------------------------------------------
    total_combo = int((events["CONTAINER_STATUS"] == "Combo").sum())
    total_twinlift = int((events["TWINLIFT_STATUS"] == "Twinlift").sum())
    total_combo_bukan_twinlift = total_combo - total_twinlift
    pct_twinlift_of_total = (total_twinlift / total_event) if total_event else 0
    pct_twinlift_of_combo = (total_twinlift / total_combo) if total_combo else 0

    dual_load = int(((out_df["STATUS"] == "Dual Cycle") & (out_df["ACTIVITY"] == "LOAD")).sum())
    dual_disc = int(((out_df["STATUS"] == "Dual Cycle") & (out_df["ACTIVITY"] == "DISC")).sum())
    single_load = int(((out_df["STATUS"] == "Non Dual") & (out_df["ACTIVITY"] == "LOAD")).sum())
    single_disc = int(((out_df["STATUS"] == "Non Dual") & (out_df["ACTIVITY"] == "DISC")).sum())

    container_load = dual_load + single_load
    container_disc = dual_disc + single_disc
    container_total = len(out_df)

    ev = events.copy()
    ev["BULAN"] = ev["START_TS"].dt.to_period("M")

    # ------------------------------------------------------
    # Breakdown bulanan -- 3 view terpisah, semuanya per event/ritase:
    #   1) Dual Cycle vs Non Dual (basis persis spt VBA)
    #   2) Combo vs Single (basis CONTAINER_STATUS, independen dari
    #      status Dual Cycle -- ini tambahan di luar VBA, sesuai
    #      permintaan, supaya kelihatan komposisi Combo/Single tiap
    #      bulan juga)
    #   3) Twinlift vs Bukan Twinlift (sub-analisis di dalam Combo,
    #      tambahan sesuai permintaan)
    # Semuanya disediakan sbg jumlah (dipakai internal) & persen
    # (dipakai buat ditampilkan, krn user minta lihat persen).
    # ------------------------------------------------------
    monthly = ev.groupby("BULAN").agg(
        total_event=("STATUS", "count"),
        dual=("STATUS", lambda s: int((s == "Dual Cycle").sum())),
        combo=("CONTAINER_STATUS", lambda s: int((s == "Combo").sum())),
        twinlift=("TWINLIFT_STATUS", lambda s: int((s == "Twinlift").sum())),
    )
    monthly["non_dual"] = monthly["total_event"] - monthly["dual"]
    monthly["single"] = monthly["total_event"] - monthly["combo"]
    monthly["combo_bukan_twinlift"] = monthly["combo"] - monthly["twinlift"]

    monthly["pct_dual"] = np.where(monthly["total_event"] > 0, monthly["dual"] / monthly["total_event"], 0)
    monthly["pct_non_dual"] = np.where(monthly["total_event"] > 0, monthly["non_dual"] / monthly["total_event"], 0)
    monthly["pct_combo"] = np.where(monthly["total_event"] > 0, monthly["combo"] / monthly["total_event"], 0)
    monthly["pct_single"] = np.where(monthly["total_event"] > 0, monthly["single"] / monthly["total_event"], 0)
    monthly["pct_twinlift"] = np.where(monthly["total_event"] > 0, monthly["twinlift"] / monthly["total_event"], 0)
    monthly["pct_twinlift_of_combo"] = np.where(
        monthly["combo"] > 0, monthly["twinlift"] / monthly["combo"], 0
    )
    monthly["pct_combo_bukan_twinlift_of_combo"] = np.where(
        monthly["combo"] > 0, monthly["combo_bukan_twinlift"] / monthly["combo"], 0
    )

    monthly = monthly.sort_index()
    monthly.index = monthly.index.astype(str)

    return {
        "total_event": total_event,
        "total_dual": total_dual,
        "total_single": total_single,
        "pct_dual": (total_dual / total_event) if total_event else 0,
        "combo_dual": combo_dual,
        "combo_single": combo_single,
        "single_dual": single_dual,
        "single_single": single_single,
        "total_combo": total_combo,
        "total_twinlift": total_twinlift,
        "total_combo_bukan_twinlift": total_combo_bukan_twinlift,
        "pct_twinlift_of_total": pct_twinlift_of_total,
        "pct_twinlift_of_combo": pct_twinlift_of_combo,
        "dual_load": dual_load,
        "dual_disc": dual_disc,
        "single_load": single_load,
        "single_disc": single_disc,
        "container_load": container_load,
        "container_disc": container_disc,
        "container_total": container_total,
        "monthly": monthly,
    }


# ================================================================
# EXPORT EXCEL (Data + Ringkasan + Chart), mirip output VBA
# ================================================================

def build_excel_download(
    out_df: pd.DataFrame, summary: dict, ambang_combo: float, ambang_dual: float, ambang_twinlift: float
) -> BytesIO:
    wb = Workbook()

    # ---- Sheet Data ----
    ws_data = wb.active
    ws_data.title = "Data"

    export_df = out_df.copy()
    for c in export_df.columns:
        if pd.api.types.is_datetime64_any_dtype(export_df[c]):
            export_df[c] = export_df[c].dt.strftime("%Y-%m-%d %H:%M:%S")

    ws_data.append(list(export_df.columns))
    for row in export_df.itertuples(index=False):
        ws_data.append(list(row))
    for cell in ws_data[1]:
        cell.font = Font(bold=True)
    for col_cells in ws_data.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws_data.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    # ---- Sheet Ringkasan ----
    ws = wb.create_sheet("Ringkasan Dual Cycle")
    ws["A1"] = "RINGKASAN ANALISIS DUAL CYCLE"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"Ambang Combo: {ambang_combo:g} menit | Ambang Dual Cycle: {ambang_dual:g} menit | "
        f"Ambang Twinlift: {ambang_twinlift:g} menit | Basis perhitungan: EVENT_ID (ritase) | "
        f"Pairing dibatasi per truk (CAR_CHE_ID)"
    )
    ws["A2"].font = Font(italic=True)

    ws["A4"] = "RINGKASAN (BERBASIS EVENT_ID / RITASE)"
    ws["A4"].font = Font(bold=True)
    ws["A5"], ws["B5"] = "Total Event (Ritase)", summary["total_event"]
    ws["A6"], ws["B6"] = "Dual Cycle", summary["total_dual"]
    ws["A6"].font = Font(bold=True)
    ws["B6"].font = Font(bold=True)
    ws["A7"], ws["B7"] = "Non Dual", summary["total_single"]
    ws["A8"], ws["B8"] = "Persentase Dual Cycle", summary["pct_dual"]
    ws["B8"].number_format = "0.0%"
    ws["A8"].font = Font(bold=True)
    ws["B8"].font = Font(bold=True)

    ws["A10"] = "RINCIAN CONTAINER (BERBASIS EVENT_ID)"
    ws["A10"].font = Font(bold=True)
    ws["A11"], ws["B11"] = "Combo - Dual Cycle", summary["combo_dual"]
    ws["A12"], ws["B12"] = "Combo - Non Dual", summary["combo_single"]
    ws["A13"], ws["B13"] = "Single - Dual Cycle", summary["single_dual"]
    ws["A14"], ws["B14"] = "Single - Non Dual", summary["single_single"]

    ws["A16"] = "RINCIAN AKTIVITAS (BERBASIS BARIS LOAD/DISC -- INFO TAMBAHAN)"
    ws["A16"].font = Font(bold=True)
    ws["A17"], ws["B17"] = "Total aktivitas (baris)", len(out_df)
    ws["A18"], ws["B18"] = "Dual Cycle - LOAD", summary["dual_load"]
    ws["A19"], ws["B19"] = "Dual Cycle - DISC", summary["dual_disc"]
    ws["A20"], ws["B20"] = "Non Dual - LOAD", summary["single_load"]
    ws["A21"], ws["B21"] = "Non Dual - DISC", summary["single_disc"]
    ws["A23"], ws["B23"] = "Jumlah Container - LOAD", summary["container_load"]
    ws["A24"], ws["B24"] = "Jumlah Container - DISC", summary["container_disc"]

    # ---- RINGKASAN TWINLIFT ----
    ws["A25"] = "RINGKASAN TWINLIFT (SUB-ANALISIS DI DALAM COMBO)"
    ws["A25"].font = Font(bold=True)
    ws["A26"], ws["B26"] = "Total Combo", summary["total_combo"]
    ws["A27"], ws["B27"] = "Twinlift", summary["total_twinlift"]
    ws["A27"].font = Font(bold=True)
    ws["B27"].font = Font(bold=True)
    ws["A28"], ws["B28"] = "Combo - Bukan Twinlift", summary["total_combo_bukan_twinlift"]
    ws["A29"], ws["B29"] = "% Twinlift dari Total Event", summary["pct_twinlift_of_total"]
    ws["B29"].number_format = "0.0%"
    ws["A30"], ws["B30"] = "% Twinlift dari Total Combo", summary["pct_twinlift_of_combo"]
    ws["B30"].number_format = "0.0%"

    ws["A32"] = "BREAKDOWN BULANAN: DUAL CYCLE VS NON DUAL (EVENT / RITASE)"
    ws["A32"].font = Font(bold=True)
    header_row = 33
    headers = ["Bulan", "Total Event", "Dual Cycle", "Non Dual", "% Dual Cycle", "% Non Dual"]
    for c_idx, label in enumerate(headers, start=1):
        cell = ws.cell(header_row, c_idx, label)
        cell.font = Font(bold=True)

    monthly = summary["monthly"]
    r = header_row
    for bulan, row in monthly.iterrows():
        r += 1
        ws.cell(r, 1, str(bulan))
        ws.cell(r, 2, int(row["total_event"]))
        ws.cell(r, 3, int(row["dual"]))
        ws.cell(r, 4, int(row["non_dual"]))
        ws.cell(r, 5, float(row["pct_dual"]))
        ws.cell(r, 5).number_format = "0.0%"
        ws.cell(r, 6, float(row["pct_non_dual"]))
        ws.cell(r, 6).number_format = "0.0%"

    total_row = r + 1
    tot_evt = int(monthly["total_event"].sum()) if len(monthly) else 0
    tot_dual = int(monthly["dual"].sum()) if len(monthly) else 0
    tot_non_dual = int(monthly["non_dual"].sum()) if len(monthly) else 0
    ws.cell(total_row, 1, "Total")
    ws.cell(total_row, 2, tot_evt)
    ws.cell(total_row, 3, tot_dual)
    ws.cell(total_row, 4, tot_non_dual)
    ws.cell(total_row, 5, (tot_dual / tot_evt) if tot_evt else 0)
    ws.cell(total_row, 5).number_format = "0.0%"
    ws.cell(total_row, 6, (tot_non_dual / tot_evt) if tot_evt else 0)
    ws.cell(total_row, 6).number_format = "0.0%"
    for c in range(1, 7):
        ws.cell(total_row, c).font = Font(bold=True)

    # ---- Breakdown bulanan tambahan: Combo vs Single (%) ----
    header_row2 = total_row + 3
    ws.cell(header_row2 - 1, 1, "BREAKDOWN BULANAN: COMBO VS SINGLE (EVENT / RITASE)")
    ws.cell(header_row2 - 1, 1).font = Font(bold=True)
    headers2 = ["Bulan", "Total Event", "Combo", "Single", "% Combo", "% Single"]
    for c_idx, label in enumerate(headers2, start=1):
        cell = ws.cell(header_row2, c_idx, label)
        cell.font = Font(bold=True)

    r2 = header_row2
    for bulan, row in monthly.iterrows():
        r2 += 1
        ws.cell(r2, 1, str(bulan))
        ws.cell(r2, 2, int(row["total_event"]))
        ws.cell(r2, 3, int(row["combo"]))
        ws.cell(r2, 4, int(row["single"]))
        ws.cell(r2, 5, float(row["pct_combo"]))
        ws.cell(r2, 5).number_format = "0.0%"
        ws.cell(r2, 6, float(row["pct_single"]))
        ws.cell(r2, 6).number_format = "0.0%"

    total_row2 = r2 + 1
    tot_combo = int(monthly["combo"].sum()) if len(monthly) else 0
    tot_single = int(monthly["single"].sum()) if len(monthly) else 0
    ws.cell(total_row2, 1, "Total")
    ws.cell(total_row2, 2, tot_evt)
    ws.cell(total_row2, 3, tot_combo)
    ws.cell(total_row2, 4, tot_single)
    ws.cell(total_row2, 5, (tot_combo / tot_evt) if tot_evt else 0)
    ws.cell(total_row2, 5).number_format = "0.0%"
    ws.cell(total_row2, 6, (tot_single / tot_evt) if tot_evt else 0)
    ws.cell(total_row2, 6).number_format = "0.0%"
    for c in range(1, 7):
        ws.cell(total_row2, c).font = Font(bold=True)

    # ---- Breakdown bulanan tambahan: Twinlift vs Bukan Twinlift (dari Combo, %) ----
    header_row3 = total_row2 + 3
    ws.cell(header_row3 - 1, 1, "BREAKDOWN BULANAN: TWINLIFT VS BUKAN TWINLIFT (DARI COMBO, %)")
    ws.cell(header_row3 - 1, 1).font = Font(bold=True)
    headers3 = ["Bulan", "Total Combo", "Twinlift", "Bukan Twinlift", "% Twinlift", "% Bukan Twinlift"]
    for c_idx, label in enumerate(headers3, start=1):
        cell = ws.cell(header_row3, c_idx, label)
        cell.font = Font(bold=True)

    r3 = header_row3
    for bulan, row in monthly.iterrows():
        r3 += 1
        ws.cell(r3, 1, str(bulan))
        ws.cell(r3, 2, int(row["combo"]))
        ws.cell(r3, 3, int(row["twinlift"]))
        ws.cell(r3, 4, int(row["combo_bukan_twinlift"]))
        ws.cell(r3, 5, float(row["pct_twinlift_of_combo"]))
        ws.cell(r3, 5).number_format = "0.0%"
        ws.cell(r3, 6, float(row["pct_combo_bukan_twinlift_of_combo"]))
        ws.cell(r3, 6).number_format = "0.0%"

    total_row3 = r3 + 1
    tot_twinlift = int(monthly["twinlift"].sum()) if len(monthly) else 0
    tot_combo_bukan_twinlift = int(monthly["combo_bukan_twinlift"].sum()) if len(monthly) else 0
    ws.cell(total_row3, 1, "Total")
    ws.cell(total_row3, 2, tot_combo)
    ws.cell(total_row3, 3, tot_twinlift)
    ws.cell(total_row3, 4, tot_combo_bukan_twinlift)
    ws.cell(total_row3, 5, (tot_twinlift / tot_combo) if tot_combo else 0)
    ws.cell(total_row3, 5).number_format = "0.0%"
    ws.cell(total_row3, 6, (tot_combo_bukan_twinlift / tot_combo) if tot_combo else 0)
    ws.cell(total_row3, 6).number_format = "0.0%"
    for c in range(1, 7):
        ws.cell(total_row3, c).font = Font(bold=True)

    ws.column_dimensions["A"].width = 46
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15

    # ---- Pie chart: Dual vs Non Dual ----
    pie = PieChart()
    pie.title = "Dual Cycle vs Non Dual (Event)"
    data = Reference(ws, min_col=2, min_row=6, max_row=7)
    cats = Reference(ws, min_col=1, min_row=6, max_row=7)
    pie.add_data(data)
    pie.set_categories(cats)
    ws.add_chart(pie, "H4")

    # ---- Bar chart: Container x Status ----
    bar = BarChart()
    bar.type = "col"
    bar.title = "Container x Status (Event)"
    data = Reference(ws, min_col=2, min_row=11, max_row=14)
    cats = Reference(ws, min_col=1, min_row=11, max_row=14)
    bar.add_data(data)
    bar.set_categories(cats)
    ws.add_chart(bar, "H20")

    # ---- Pie chart: Twinlift vs Bukan Twinlift (dari Combo) ----
    pie_tw = PieChart()
    pie_tw.title = "Twinlift vs Bukan Twinlift (dari Combo)"
    data = Reference(ws, min_col=2, min_row=27, max_row=28)
    cats = Reference(ws, min_col=1, min_row=27, max_row=28)
    pie_tw.add_data(data)
    pie_tw.set_categories(cats)
    ws.add_chart(pie_tw, "H36")

    # ---- Stacked bar (%): breakdown bulanan Dual Cycle vs Non Dual ----
    if len(monthly) > 0:
        bar2 = BarChart()
        bar2.type = "col"
        bar2.grouping = "percentStacked"
        bar2.overlap = 100
        bar2.title = "Breakdown Bulanan: Dual Cycle vs Non Dual (%)"
        data = Reference(ws, min_col=3, max_col=4, min_row=header_row, max_row=total_row - 1)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=total_row - 1)
        bar2.add_data(data, titles_from_data=True)
        bar2.set_categories(cats)
        ws.add_chart(bar2, "H52")

        # ---- Stacked bar (%): breakdown bulanan Combo vs Single ----
        bar3 = BarChart()
        bar3.type = "col"
        bar3.grouping = "percentStacked"
        bar3.overlap = 100
        bar3.title = "Breakdown Bulanan: Combo vs Single (%)"
        data3 = Reference(ws, min_col=3, max_col=4, min_row=header_row2, max_row=total_row2 - 1)
        cats3 = Reference(ws, min_col=1, min_row=header_row2 + 1, max_row=total_row2 - 1)
        bar3.add_data(data3, titles_from_data=True)
        bar3.set_categories(cats3)
        ws.add_chart(bar3, "H68")

        # ---- Stacked bar (%): breakdown bulanan Twinlift vs Bukan Twinlift ----
        bar4 = BarChart()
        bar4.type = "col"
        bar4.grouping = "percentStacked"
        bar4.overlap = 100
        bar4.title = "Breakdown Bulanan: Twinlift vs Bukan Twinlift (% dari Combo)"
        data4 = Reference(ws, min_col=3, max_col=4, min_row=header_row3, max_row=total_row3 - 1)
        cats4 = Reference(ws, min_col=1, min_row=header_row3 + 1, max_row=total_row3 - 1)
        bar4.add_data(data4, titles_from_data=True)
        bar4.set_categories(cats4)
        ws.add_chart(bar4, "H84")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ================================================================
# ================================================================
# UI STREAMLIT
# ================================================================
# ================================================================

# ----------------------------------------------------------------
# Path logo dihitung RELATIF terhadap lokasi file app.py ini sendiri
# (bukan relatif terhadap current working directory), supaya tetap
# ketemu walau di-deploy di Streamlit Cloud yang CWD-nya bisa beda
# dari folder script. File logo harus ada di:
#   <folder app.py>/assets/logo_caca_icon.png  (ikon CACA saja, tanpa teks)
#   <folder app.py>/assets/logo_pelindo.png
# ----------------------------------------------------------------
ASSETS_DIR = Path(__file__).resolve().parent / "assets"
LOGO_CACA_ICON_PATH = ASSETS_DIR / "logo_caca_icon.png"
LOGO_PATH = ASSETS_DIR / "logo_pelindo.png"


def _img_to_base64(path: Path) -> str:
    import base64
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def _render_header(icon_path: Path, brand_path: Path):
    """
    Header custom pakai flexbox HTML (bukan st.columns) supaya ikon,
    judul "CACA", subjudul, dan logo Pelindo di kanan bisa presisi
    sejajar vertikal (align-items: center) mirip layout referensi.
    """
    icon_ok = icon_path.exists()
    brand_ok = brand_path.exists()

    icon_html = (
        f'<img src="data:image/png;base64,{_img_to_base64(icon_path)}" '
        f'style="height:44px;width:auto;display:block;" />'
        if icon_ok else ""
    )
    brand_html = (
        f'<img src="data:image/png;base64,{_img_to_base64(brand_path)}" '
        f'style="height:40px;width:auto;display:block;" />'
        if brand_ok else ""
    )

    st.markdown(
        f"""
        <div style="display:flex;align-items:center;justify-content:space-between;
                    padding:6px 0 2px 0;">
            <div style="display:flex;align-items:center;gap:12px;">
                {icon_html}
                <div style="line-height:1.15;">
                    <div style="font-size:1.55rem;font-weight:800;color:#16324f;
                                letter-spacing:0.5px;">CACA</div>
                    <div style="font-size:0.72rem;color:#6b7280;">
                        Cycle Analysis and Cargo Optimalization
                    </div>
                </div>
            </div>
            <div>{brand_html}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not icon_ok:
        st.caption(f"⚠️ Logo CACA tidak ditemukan di `{icon_path}`.")
    if not brand_ok:
        st.caption(f"⚠️ Logo Pelindo tidak ditemukan di `{brand_path}`.")


_render_header(LOGO_CACA_ICON_PATH, LOGO_PATH)
st.divider()

# ----------------------------------------------------------------
# PENGATURAN -- ditaruh langsung di dashboard (bukan sidebar) biar
# layout terasa lebih lega/luas. Ukuran kontainer eligible utk Combo
# TIDAK lagi jadi pengaturan yang bisa diubah -- dikunci 20ft persis
# spt VBA (lihat SIZE_ELIGIBLE di atas), karena mengubahnya bisa bikin
# hasil beda dari VBA.
# ----------------------------------------------------------------
size_eligible = SIZE_ELIGIBLE

with st.expander("⚙️ Pengaturan Ambang Batas Analisis", expanded=False):
    pc1, pc2, pc3 = st.columns(3)
    with pc1:
        ambang_combo = st.number_input(
            "Ambang Combo (menit)", min_value=1, value=AMBANG_COMBO_MENIT_DEFAULT, step=5,
            help="Jarak waktu maksimum antar 2 baris size 20ft, truk & aktivitas sama, supaya dianggap 'Combo'.",
        )
    with pc2:
        ambang_dual = st.number_input(
            "Ambang Dual Cycle (menit)", min_value=1, value=AMBANG_DUAL_MENIT_DEFAULT, step=10,
            help="Jarak waktu maksimum antar 2 event beda aktivitas (LOAD vs DISC) dalam truk yang sama.",
        )
    with pc3:
        ambang_twinlift = st.number_input(
            "Ambang Twinlift (menit)", min_value=1, value=AMBANG_TWINLIFT_MENIT_DEFAULT, step=1,
            help=(
                "Jarak waktu maksimum DISC_LOAD_TS antar 2 kontainer dalam 1 Combo 20ft, "
                "yang berasal dari kapal (VES_ID) & truk yang sama, supaya dianggap 'Twinlift'."
            ),
        )
    st.caption(f"Ukuran kontainer eligible untuk Combo/Twinlift dikunci **{SIZE_ELIGIBLE} ft**, mengikuti macro VBA.")

st.subheader("1️⃣ Upload Data")
uploaded = st.file_uploader("Upload file .xlsx atau .csv", type=["xlsx", "xls", "csv"])

if uploaded is None:
    st.session_state.pop("hasil", None)
    st.session_state.pop("_last_file_sig", None)
    st.info("Silakan upload file data untuk memulai analisis.")
    st.stop()

file_bytes = uploaded.getvalue()

# --------------------------------------------------------------
# Deteksi kalau file yang diupload BERBEDA dari sebelumnya
# (nama, ukuran, atau isinya beda) -> otomatis bersihkan hasil
# analisis lama supaya tidak nyangkut / bikin bingung. Ini juga
# yang bikin ganti-ganti file jadi tidak perlu reboot app lagi.
# --------------------------------------------------------------
file_sig = (uploaded.name, len(file_bytes), hash(file_bytes[:1_000_000]))
if st.session_state.get("_last_file_sig") != file_sig:
    st.session_state.pop("hasil", None)
    st.session_state["_last_file_sig"] = file_sig

try:
    sheets = baca_file(file_bytes, uploaded.name)
except Exception as e:
    st.error(
        "❌ Gagal membaca file yang diupload. Pastikan file tidak corrupt dan "
        "formatnya benar-benar .xlsx / .xls / .csv."
    )
    with st.expander("Detail error (untuk dilaporkan)"):
        st.exception(e)
    st.stop()

if not sheets:
    st.error("❌ File tidak berisi sheet/data apa pun.")
    st.stop()

sheet_name = None
if len(sheets) > 1:
    sheet_name = st.selectbox("Pilih sheet data", list(sheets.keys()))
else:
    sheet_name = list(sheets.keys())[0]

raw = sheets[sheet_name]
st.caption(f"File terbaca: {len(raw)} baris data.")

st.subheader("2️⃣ Pemetaan Kolom")
cols = list(raw.columns)


def guess(options, keywords, default_idx=0):
    for kw in keywords:
        for i, c in enumerate(options):
            if kw.lower() in str(c).lower():
                return i
    return default_idx


c1, c2, c3 = st.columns(3)
with c1:
    col_ves = st.selectbox("VES_ID", cols, index=guess(cols, ["ves"]))
    col_size = st.selectbox("CTR_SIZE (ukuran kontainer)", cols, index=guess(cols, ["size", "ctr_size"]))
with c2:
    col_truck = st.selectbox("CAR_CHE_ID (truk)", cols, index=guess(cols, ["car_che", "truck", "che"]))
    col_activity = st.selectbox("ACTIVITY", cols, index=guess(cols, ["activity", "aktivitas"]))
with c3:
    col_ts_g = st.selectbox("DISC_LOAD_TS", cols, index=guess(cols, ["disc_load", "disc_loading"]))
    col_ts_h = st.selectbox("STACK_UNSTACK_TS", cols, index=guess(cols, ["stack_unstack", "unstack_stack"]))

col_map = {
    "ves_id": col_ves,
    "size": col_size,
    "truck": col_truck,
    "activity": col_activity,
    "ts_g": col_ts_g,
    "ts_h": col_ts_h,
}

run = st.button("▶️ Jalankan Analisis Dual Cycle", type="primary")

if not run and "hasil" not in st.session_state:
    st.info("👆 Klik tombol di atas untuk menjalankan analisis pada file ini.")
    st.stop()

if run:
    try:
        with st.spinner("Memproses data..."):
            df = siapkan_data(raw, col_map, size_eligible)

            if len(df) == 0:
                st.error(
                    "❌ Setelah pembersihan, tidak ada baris data yang tersisa. "
                    "Kemungkinan kolom DISC_LOAD_TS & STACK_UNSTACK_TS yang dipilih "
                    "tidak berisi tanggal/jam yang valid, atau pemetaan kolom di "
                    "Langkah 2 salah. Silakan periksa kembali pemetaan kolom di atas."
                )
                st.stop()

            df_combo = layer1_combo(df, ambang_combo, size_eligible)
            twinlift_map = deteksi_twinlift(df_combo, ambang_twinlift, size_eligible)
            events = bentuk_event(df_combo)
            events["TWINLIFT_STATUS"] = events["GROUP_ID"].map(twinlift_map)
            events = layer2_dual(events, ambang_dual)
            events, event_id_map = beri_event_id(events, df_combo)
            out_df = gabungkan_hasil(df_combo, events, event_id_map)
            summary = hitung_ringkasan(events, out_df)
    except Exception as e:
        st.error(
            "❌ Terjadi error saat memproses data. Detail error ada di bawah ini — "
            "silakan screenshot/copy untuk dilaporkan."
        )
        with st.expander("Detail error (untuk dilaporkan)", expanded=True):
            st.exception(e)
        st.stop()

    st.session_state["hasil"] = {
        "out_df": out_df,
        "events": events,
        "summary": summary,
        "ambang_combo": ambang_combo,
        "ambang_dual": ambang_dual,
        "ambang_twinlift": ambang_twinlift,
    }

hasil = st.session_state["hasil"]
out_df = hasil["out_df"]
events = hasil["events"]
summary = hasil["summary"]

st.success("Analisis selesai!")

monthly = summary["monthly"].reset_index().rename(columns={"BULAN": "Bulan"})

# ================================================================
# DASHBOARD HASIL -- ditampilkan sbg 3 menu/tab: Dual Cycle,
# Twinlift, dan Download Hasil Analisis, langsung berisi grafik
# tanpa perlu scroll panjang lewat banyak subheader.
# ================================================================

tab_dual, tab_twinlift, tab_download = st.tabs(
    ["📊 Dual Cycle", "🔗 Twinlift", "⬇️ Download Hasil Analisis"]
)

# ----------------------------------------------------------------
# TAB 1: DUAL CYCLE
# ----------------------------------------------------------------
with tab_dual:
    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Total Event (Ritase)", summary["total_event"])
    k2.metric("Dual Cycle", summary["total_dual"])
    k3.metric("Non Dual", summary["total_single"])
    k4.metric("% Dual Cycle", f"{summary['pct_dual']*100:.1f}%")
    k5.metric("Jumlah Container - LOAD", summary["container_load"])
    k6.metric("Jumlah Container - DISC", summary["container_disc"])

    cc1, cc2 = st.columns(2)

    with cc1:
        pie_df = pd.DataFrame(
            {"Status": ["Dual Cycle", "Non Dual"], "Jumlah": [summary["total_dual"], summary["total_single"]]}
        )
        fig_pie = px.pie(
            pie_df, names="Status", values="Jumlah", hole=0.45,
            title="Dual Cycle vs Non Dual (berbasis Event)",
            color="Status",
            color_discrete_map={"Dual Cycle": "#2E86AB", "Non Dual": "#E76F51"},
        )
        fig_pie.update_traces(textinfo="percent+label")
        st.plotly_chart(fig_pie, width='stretch')

    with cc2:
        container_df = pd.DataFrame(
            {
                "Container": ["Combo", "Combo", "Single", "Single"],
                "Status": ["Dual Cycle", "Non Dual", "Dual Cycle", "Non Dual"],
                "Jumlah": [
                    summary["combo_dual"], summary["combo_single"],
                    summary["single_dual"], summary["single_single"],
                ],
            }
        )
        fig_bar = px.bar(
            container_df, x="Container", y="Jumlah", color="Status", barmode="group",
            title="Rincian Container x Status (Event)",
            color_discrete_map={"Dual Cycle": "#2E86AB", "Non Dual": "#E76F51"},
            text="Jumlah",
        )
        st.plotly_chart(fig_bar, width='stretch')

    if len(monthly) > 0:
        # Breakdown bulanan Dual Cycle vs Non Dual (%)
        monthly_dual_pct = monthly.melt(
            id_vars="Bulan",
            value_vars=["pct_dual", "pct_non_dual"],
            var_name="Kategori",
            value_name="Persentase",
        )
        monthly_dual_pct["Kategori"] = monthly_dual_pct["Kategori"].map(
            {"pct_dual": "Dual Cycle", "pct_non_dual": "Non Dual"}
        )
        monthly_dual_pct["Persentase"] = monthly_dual_pct["Persentase"] * 100

        fig_month_dual = px.bar(
            monthly_dual_pct, x="Bulan", y="Persentase", color="Kategori", barmode="stack",
            title="Breakdown Bulanan: Dual Cycle vs Non Dual (%)",
            color_discrete_map={"Dual Cycle": "#2E86AB", "Non Dual": "#E76F51"},
            text_auto=".1f",
        )
        fig_month_dual.update_layout(yaxis=dict(title="% dari Total Event", range=[0, 100]))
        st.plotly_chart(fig_month_dual, width='stretch')

        # Breakdown bulanan Combo vs Single (%)
        monthly_container_pct = monthly.melt(
            id_vars="Bulan",
            value_vars=["pct_combo", "pct_single"],
            var_name="Kategori",
            value_name="Persentase",
        )
        monthly_container_pct["Kategori"] = monthly_container_pct["Kategori"].map(
            {"pct_combo": "Combo", "pct_single": "Single"}
        )
        monthly_container_pct["Persentase"] = monthly_container_pct["Persentase"] * 100

        fig_month_container = px.bar(
            monthly_container_pct, x="Bulan", y="Persentase", color="Kategori", barmode="stack",
            title="Breakdown Bulanan: Combo vs Single (%)",
            color_discrete_map={"Combo": "#F4A261", "Single": "#8AB17D"},
            text_auto=".1f",
        )
        fig_month_container.update_layout(yaxis=dict(title="% dari Total Event", range=[0, 100]))
        st.plotly_chart(fig_month_container, width='stretch')

        with st.expander("📋 Tabel Breakdown Bulanan (dalam persen)"):
            tabel_bulan = monthly[
                ["Bulan", "total_event", "pct_dual", "pct_non_dual", "pct_combo", "pct_single"]
            ].copy()
            tabel_bulan.columns = [
                "Bulan", "Total Event", "% Dual Cycle", "% Non Dual", "% Combo", "% Single",
            ]
            for c in ["% Dual Cycle", "% Non Dual", "% Combo", "% Single"]:
                tabel_bulan[c] = (tabel_bulan[c] * 100).round(1).astype(str) + "%"
            st.dataframe(tabel_bulan, width='stretch', hide_index=True)

    with st.expander("📋 Rincian Aktivitas (berbasis baris LOAD/DISC — info tambahan)"):
        act_df = pd.DataFrame(
            {
                "Status": ["Dual Cycle", "Dual Cycle", "Non Dual", "Non Dual"],
                "Activity": ["LOAD", "DISC", "LOAD", "DISC"],
                "Jumlah": [
                    summary["dual_load"], summary["dual_disc"],
                    summary["single_load"], summary["single_disc"],
                ],
            }
        )
        st.dataframe(act_df, width='stretch', hide_index=True)

# ----------------------------------------------------------------
# TAB 2: TWINLIFT
# ----------------------------------------------------------------
with tab_twinlift:
    st.caption(
        "Twinlift = kontainer di dalam Combo 20ft, berasal dari kapal (VES_ID) & truk yang sama, "
        f"dengan jarak DISC_LOAD_TS antar keduanya ≤ {hasil['ambang_twinlift']:g} menit."
    )

    t1, t2, t3, t4, t5 = st.columns(5)
    t1.metric("Total Combo", summary["total_combo"])
    t2.metric("Twinlift", summary["total_twinlift"])
    t3.metric("Combo - Bukan Twinlift", summary["total_combo_bukan_twinlift"])
    t4.metric("% Twinlift dari Combo", f"{summary['pct_twinlift_of_combo']*100:.1f}%")
    t5.metric("% Twinlift dari Total Event", f"{summary['pct_twinlift_of_total']*100:.1f}%")

    tc1, tc2 = st.columns(2)

    with tc1:
        if summary["total_combo"] > 0:
            twin_df = pd.DataFrame(
                {
                    "Status": ["Twinlift", "Bukan Twinlift"],
                    "Jumlah": [summary["total_twinlift"], summary["total_combo_bukan_twinlift"]],
                }
            )
            fig_twin_pie = px.pie(
                twin_df, names="Status", values="Jumlah", hole=0.45,
                title="Twinlift vs Bukan Twinlift (dari Combo)",
                color="Status",
                color_discrete_map={"Twinlift": "#6A4C93", "Bukan Twinlift": "#C0C0C0"},
            )
            fig_twin_pie.update_traces(textinfo="percent+label")
            st.plotly_chart(fig_twin_pie, width='stretch')
        else:
            st.info("Tidak ada event Combo pada data ini, sehingga tidak ada kandidat Twinlift.")

    with tc2:
        if len(monthly) > 0 and summary["total_combo"] > 0:
            monthly_twin_pct = monthly.melt(
                id_vars="Bulan",
                value_vars=["pct_twinlift_of_combo", "pct_combo_bukan_twinlift_of_combo"],
                var_name="Kategori",
                value_name="Persentase",
            )
            monthly_twin_pct["Kategori"] = monthly_twin_pct["Kategori"].map(
                {"pct_twinlift_of_combo": "Twinlift", "pct_combo_bukan_twinlift_of_combo": "Bukan Twinlift"}
            )
            monthly_twin_pct["Persentase"] = monthly_twin_pct["Persentase"] * 100

            fig_month_twin = px.bar(
                monthly_twin_pct, x="Bulan", y="Persentase", color="Kategori", barmode="stack",
                title="Breakdown Bulanan: Twinlift vs Bukan Twinlift (% dari Combo)",
                color_discrete_map={"Twinlift": "#6A4C93", "Bukan Twinlift": "#C0C0C0"},
                text_auto=".1f",
            )
            fig_month_twin.update_layout(yaxis=dict(title="% dari Total Combo", range=[0, 100]))
            st.plotly_chart(fig_month_twin, width='stretch')

    if len(monthly) > 0:
        with st.expander("📋 Tabel Breakdown Bulanan Twinlift (dari Combo, dalam persen)"):
            tabel_twin = monthly[
                ["Bulan", "combo", "twinlift", "combo_bukan_twinlift", "pct_twinlift_of_combo",
                 "pct_combo_bukan_twinlift_of_combo"]
            ].copy()
            tabel_twin.columns = [
                "Bulan", "Total Combo", "Twinlift", "Bukan Twinlift", "% Twinlift", "% Bukan Twinlift",
            ]
            for c in ["% Twinlift", "% Bukan Twinlift"]:
                tabel_twin[c] = (tabel_twin[c] * 100).round(1).astype(str) + "%"
            st.dataframe(tabel_twin, width='stretch', hide_index=True)

# ----------------------------------------------------------------
# TAB 3: DOWNLOAD HASIL ANALISIS
# ----------------------------------------------------------------
with tab_download:
    st.write(f"Data hasil analisis ({len(out_df)} baris):")
    st.dataframe(out_df, width='stretch', height=400)

    try:
        excel_bio = build_excel_download(
            out_df, summary, hasil["ambang_combo"], hasil["ambang_dual"], hasil["ambang_twinlift"]
        )
    except Exception as e:
        st.error(
            "❌ Gagal membuat file Excel hasil analisis. Data hasil tetap bisa dilihat "
            "di tabel & di-download sebagai CSV di bawah."
        )
        with st.expander("Detail error (untuk dilaporkan)"):
            st.exception(e)
        excel_bio = None

    dcol1, dcol2 = st.columns(2)
    with dcol1:
        st.download_button(
            "⬇️ Download Excel (Data + Ringkasan + Chart)",
            data=excel_bio if excel_bio is not None else b"",
            file_name="Hasil_Analisis_Dual_Cycle.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch',
            disabled=excel_bio is None,
        )
    with dcol2:
        csv_bytes = out_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Download CSV (Data saja)",
            data=csv_bytes,
            file_name="Hasil_Analisis_Dual_Cycle.csv",
            mime="text/csv",
            width='stretch',
        )
